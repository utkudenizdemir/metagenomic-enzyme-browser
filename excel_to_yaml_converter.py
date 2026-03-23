# excel_to_yaml_converter.py
# ============================================================
# This Python program reads data from an Excel spreadsheet 
# and converts it into a YAML file for use in database population
# or other applications. The spreadsheet must contain 
# multiple sheets corresponding to the tables in your database
# schema (e.g., Study, Biome, Sample, etc.).
#
# The program will:
# 1. Read data from an Excel file.
# 2. Extract information from predefined sheets and structure it.
# 3. Generate a YAML file with the corresponding data.
#
# This program assumes the following sheet structure:
# 1. 'Study' sheet should contain study data.
# 2. 'Biome' sheet should contain biome data.
# 3. 'Sample' sheet should contain sample data.
# 4. 'Geolocation' sheet should contain geolocation data.
# 5. 'DNAPrep' sheet should contain DNA prep data.
# 6. 'SequencingRun' sheet should contain sequencing run data.
# 7. 'Assembly' sheet should contain assembly data.
# 8. 'OverallStats' sheet should contain overall statistics data.
#
# It will generate a YAML file with the same structure as the
# given YAML example.
#
# Usage Instructions:
# --------------------
# 1. Make sure you have the required Python packages installed:
#    - openpyxl (to handle Excel files)
#    - PyYAML (to write YAML files)
#
# 2. You can install the required packages using pip:
#    pip install openpyxl pyyaml
#
# 3. Place your Excel file in the same directory as this script,
#    or update the script to specify the correct path.
#
# 4. Run the script using:
#    python excel_to_yaml_converter.py <input_excel_file.xlsx> <output_yaml_file.yaml>
#
# 5. The script will generate a YAML file based on the input Excel file.
#
# Note: Ensure that your Excel file follows the correct format,
#       with column names matching those expected for each sheet.
# ============================================================

import openpyxl
import yaml
import sys
import os

class ExcelToYamlConverter:
    def __init__(self, excel_file, yaml_file):
        """
        Initialize the converter with input and output file paths.
        
        :param excel_file: Path to the input Excel file
        :param yaml_file: Path to the output YAML file
        """
        self.excel_file = excel_file
        self.yaml_file = yaml_file
        self.data = {
            "Study": [],
            "Biome": [],
            "Sample": [],
            "Geolocation": [],
            "DNAPrep": [],
            "SequencingRun": [],
            "Assembly": [],
            "OverallStats": []
        }

    def load_excel_data(self):
        """
        Load data from the provided Excel file into the class's data attribute.
        Each sheet in the Excel file should be named after the corresponding table.
        """
        try:
            # Load the Excel workbook
            wb = openpyxl.load_workbook(self.excel_file)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                if sheet_name in self.data:
                    self.data[sheet_name] = self._parse_sheet(sheet)
                else:
                    print(f"Warning: Sheet '{sheet_name}' is not recognized. Skipping.")
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            sys.exit(1)

    def _parse_sheet(self, sheet):
        """
        Parse a sheet into a list of dictionaries, where each row corresponds to a dictionary
        with column names as keys.

        :param sheet: The openpyxl sheet object to parse
        :return: List of dictionaries representing rows in the sheet
        """
        rows = []
        headers = [cell.value for cell in sheet[1]]  # First row as header
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {headers[i]: row[i] for i in range(len(headers))}
            rows.append(row_data)
        return rows

    def save_yaml(self):
        """
        Save the collected data into a YAML file.
        """
        try:
            with open(self.yaml_file, 'w') as file:
                yaml.dump(self.data, file, default_flow_style=False)
            print(f"YAML file has been saved to {self.yaml_file}")
        except Exception as e:
            print(f"Error saving YAML file: {e}")
            sys.exit(1)

    def convert(self):
        """
        Perform the full conversion process:
        1. Load data from Excel.
        2. Save the corresponding YAML file.
        """
        self.load_excel_data()
        self.save_yaml()

def main():
    if len(sys.argv) != 3:
        print("Usage: python excel_to_yaml_converter.py <input_excel_file.xlsx> <output_yaml_file.yaml>")
        sys.exit(1)

    excel_file = sys.argv[1]
    yaml_file = sys.argv[2]

    if not os.path.exists(excel_file):
        print(f"Error: The file '{excel_file}' does not exist.")
        sys.exit(1)

    converter = ExcelToYamlConverter(excel_file, yaml_file)
    converter.convert()

if __name__ == "__main__":
    main()
